import docx
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def convert_word_to_excel(word_file, excel_file):
    doc = docx.Document(word_file)
    wb = Workbook()
    ws = wb.active

    # セルのデフォルトの配置を左上に設定
    align = Alignment(vertical="top", horizontal="left")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = align

    page_number = 1
    for i, paragraph in enumerate(doc.paragraphs):
        # 空の段落をスキップ
        if not paragraph.text.strip():
            continue

        # ページごとに新しいシートを作成
        if i != 0 and i % 50 == 0:
            ws = wb.create_sheet(title=f"Page {page_number}")
            page_number += 1

        # 改行に基づいて段落を行ごとに分割
        lines = paragraph.text.split("\n")
        row_index = len(ws["A"]) + 1

        # 各行を別のセルに書き込む
        for j, line in enumerate(lines):
            ws.cell(row=row_index + j, column=1).value = line

        # もし段落が複数行ある場合はセルを横方向に結合
        if len(lines) > 1:
            merge_range = f"A{row_index}:A{row_index + len(lines) - 1}"
            ws.merge_cells(merge_range)

    # 列幅を自動調整
    for column in ws.columns:
        max_length = 0
        for cell in column:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except TypeError:
                    pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Excelファイルを保存
    wb.save(excel_file)


# 使用例
word_file = "assets/記録書テスト用_変換前.doc"
excel_file = "assets/記録書テスト用_変換後.xlsx"
convert_word_to_excel(word_file, excel_file)
